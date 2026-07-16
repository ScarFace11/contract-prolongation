# =============================================================================
# Анализ пролонгаций аккаунт-менеджеров за 2023 год
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# 1. ЗАГРУЗКА ДАННЫХ
# =============================================================================

prolongations = pd.read_csv(
    "attached_assets/prolongations_1784238197480.csv",
    sep=None, engine="python", encoding="utf-8-sig"
)
financial = pd.read_csv(
    "attached_assets/financial_data_1784238197480.csv",
    sep=None, engine="python", encoding="utf-8-sig"
)

# =============================================================================
# 2. ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =============================================================================

months_column = {
    1: "Январь", 2: "Февраль", 3: "Март",    4: "Апрель",
    5: "Май",    6: "Июнь",    7: "Июль",    8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}
months_parse = {v.lower(): k for k, v in months_column.items()}

def parse_month(value):
    """'январь 2023' → pd.Timestamp"""
    month_str, year_str = value.lower().split()
    return pd.Timestamp(year=int(year_str), month=months_parse[month_str], day=1)

def month_label(date):
    """pd.Timestamp → 'Январь 2023' (имя колонки в financial)"""
    return f"{months_column[date.month]} {date.year}"

def money_to_float(value):
    """'36 220,00' / 'стоп' / NaN → float"""
    if pd.isna(value):
        return 0.0
    if isinstance(value, str):
        cleaned = value.replace("\xa0", "").replace(" ", "").replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return float(value)

def fmt_pct(x):
    """float → '44.7%'; NaN → 'Нет данных'"""
    if pd.isna(x):
        return "Нет данных"
    return f"{x:.1%}"

def fmt_money(x):
    """float → '5 986 767'"""
    if pd.isna(x) or x == 0:
        return "0"
    return f"{x:,.0f}".replace(",", " ")

# =============================================================================
# 3. ПОДГОТОВКА ДАННЫХ
# =============================================================================

prolongations["month"] = prolongations["month"].apply(parse_month)
financial["id"]        = financial["id"].astype(str)
prolongations["id"]    = prolongations["id"].astype(str)

service_columns = ["id", "Причина дубля", "Account"]
month_columns   = [c for c in financial.columns if c not in service_columns]

for col in month_columns:
    financial[col] = financial[col].apply(money_to_float)

# Агрегируем дубли (первая/вторая часть оплаты) по id
financial_grouped = (
    financial.groupby("id")[month_columns].sum().reset_index()
)

# Дедублицируем prolongations по id+month (оставляем первую запись).
# В данных есть технические дубли (одинаковый AM дважды: id 600, 682, 697)
# и один конфликт AM (id 361: Смирнова vs Попова → берём первую запись, Смирнову).
prolongations = prolongations.drop_duplicates(subset=["id", "month"], keep="first")

# Объединяем с prolongations (AM из prolongations — первичен)
projects = (
    prolongations[["id", "month", "AM"]]
    .merge(financial_grouped, on="id", how="left")
)
projects["AM"] = projects["AM"].fillna("Не указан")
projects[month_columns] = projects[month_columns].fillna(0)

# =============================================================================
# 4. РАСЧЁТ KPI ПРОЛОНГАЦИЙ
# =============================================================================

def calculate_manager_kpi(data, report_month):
    """
    Коэффициент 1 месяц (пример — май 2023):
      Знаменатель — сумма отгрузки за АПРЕЛЬ проектов, завершившихся в апреле
      Числитель   — сумма отгрузки за МАЙ тех же проектов

    Коэффициент 2 месяц (пример — май 2023):
      Берём проекты, завершившиеся в МАРТЕ и не продлённые в апреле (апрель = 0)
      Знаменатель — сумма отгрузки за МАРТ таких проектов
      Числитель   — сумма отгрузки за МАЙ таких проектов
    """
    result = []
    current   = pd.Timestamp(report_month)
    previous  = current - pd.DateOffset(months=1)
    two_month = current - pd.DateOffset(months=2)

    current_col   = month_label(current)
    previous_col  = month_label(previous)
    two_month_col = month_label(two_month)

    for col in [current_col, previous_col, two_month_col]:
        if col not in month_columns:
            return pd.DataFrame()

    for manager, df in data.groupby("AM"):
        # ── Коэффициент 1 ─────────────────────────────────────────────────
        first_proj   = df[df["month"] == previous]
        denom1       = first_proj[previous_col].sum()
        numer1       = first_proj[current_col].sum()
        # np.nan если у менеджера не было завершившихся проектов в этом месяце
        # (0.0 означало бы «пролонгировал 0%» — неверный сигнал для руководителя)
        coef1        = numer1 / denom1 if denom1 > 0 else np.nan

        # ── Коэффициент 2 ─────────────────────────────────────────────────
        second_proj  = df[df["month"] == two_month]
        not_extended = second_proj[previous_col] <= 0   # не продлились в previous
        denom_proj   = second_proj[not_extended]
        denom2       = denom_proj[two_month_col].sum()
        numer2       = denom_proj[current_col].sum()
        # np.nan если нет непродлённых проектов из позапрошлого месяца
        coef2        = numer2 / denom2 if denom2 > 0 else np.nan

        result.append({
            "Менеджер":            manager,
            "Месяц":               current.strftime("%Y-%m"),
            "к пролонгации 1":     denom1,
            "пролонгировано 1":    numer1,
            "Коэффициент 1":       coef1,
            "к пролонгации 2":     denom2,
            "пролонгировано 2":    numer2,
            "Коэффициент 2":       coef2,
        })
    return pd.DataFrame(result)

all_results = []
for month in pd.date_range("2023-01-01", "2023-12-01", freq="MS"):
    temp = calculate_manager_kpi(projects, month)
    if not temp.empty:
        all_results.append(temp)

kpi_monthly = pd.concat(all_results, ignore_index=True)

# =============================================================================
# 5. АГРЕГАЦИЯ
# =============================================================================

# Весь отдел по месяцам
department = (
    kpi_monthly.groupby("Месяц")
    .agg({"к пролонгации 1":"sum","пролонгировано 1":"sum",
          "к пролонгации 2":"sum","пролонгировано 2":"sum"})
    .reset_index()
)
department["Коэффициент 1"] = np.where(
    department["к пролонгации 1"] > 0,
    department["пролонгировано 1"] / department["к пролонгации 1"], 0.0)
department["Коэффициент 2"] = np.where(
    department["к пролонгации 2"] > 0,
    department["пролонгировано 2"] / department["к пролонгации 2"], 0.0)

# Менеджеры за год
year_manager = (
    kpi_monthly.groupby("Менеджер")
    .agg({"к пролонгации 1":"sum","пролонгировано 1":"sum",
          "к пролонгации 2":"sum","пролонгировано 2":"sum"})
    .reset_index()
)
year_manager["Коэффициент 1"] = np.where(
    year_manager["к пролонгации 1"] > 0,
    year_manager["пролонгировано 1"] / year_manager["к пролонгации 1"], np.nan)
year_manager["Коэффициент 2"] = np.where(
    year_manager["к пролонгации 2"] > 0,
    year_manager["пролонгировано 2"] / year_manager["к пролонгации 2"], np.nan)

# Весь отдел за год (одна итоговая строка)
dept_k1_sum = department["к пролонгации 1"].sum()
dept_p1_sum = department["пролонгировано 1"].sum()
dept_k2_sum = department["к пролонгации 2"].sum()
dept_p2_sum = department["пролонгировано 2"].sum()
department_year = {
    "к пролонгации 1":  dept_k1_sum,
    "пролонгировано 1": dept_p1_sum,
    "Коэффициент 1":    dept_p1_sum / dept_k1_sum if dept_k1_sum > 0 else 0.0,
    "к пролонгации 2":  dept_k2_sum,
    "пролонгировано 2": dept_p2_sum,
    "Коэффициент 2":    dept_p2_sum / dept_k2_sum if dept_k2_sum > 0 else 0.0,
}

# Pivot: менеджер × месяц для каждого коэффициента
pivot_coef1 = (
    kpi_monthly.pivot_table(
        index="Менеджер", columns="Месяц", values="Коэффициент 1", aggfunc="mean")
    .reset_index()
)
pivot_coef2 = (
    kpi_monthly.pivot_table(
        index="Менеджер", columns="Месяц", values="Коэффициент 2", aggfunc="mean")
    .reset_index()
)

# =============================================================================
# 6. ВИЗУАЛИЗАЦИЯ (числовые данные — до форматирования в строки)
# =============================================================================

fig, axes = plt.subplots(1, 2, figsize=(16, 5))
fig.suptitle("Динамика коэффициентов пролонгации отдела — 2023",
             fontsize=14, fontweight="bold")

for ax, col, color, title in [
    (axes[0], "Коэффициент 1", "#F4A800", "Коэффициент пролонгации: 1-й месяц"),
    (axes[1], "Коэффициент 2", "#E07B00", "Коэффициент пролонгации: 2-й месяц"),
]:
    ax.plot(department["Месяц"], department[col], marker="o", color=color)
    ax.set_title(title)
    ax.set_xlabel("Месяц")
    ax.set_ylabel("Коэффициент")
    ax.yaxis.set_major_formatter(mticker.PercentFormatter(xmax=1))
    ax.tick_params(axis="x", rotation=45)
    ax.grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig("prolongation_chart.png", dpi=150, bbox_inches="tight")
plt.close()
print("График сохранён: prolongation_chart.png")

# =============================================================================
# 7. EXCEL — строго по шаблону ТЗ
# =============================================================================

# Цвета из шаблона (жёлто-бежевая гамма)
C_GROUP  = "FFD966"   # тёмно-жёлтый — строка с названием группы ("Пролонгации в первый месяц")
C_HEADER = "FFF2CC"   # светло-жёлтый — строка с подзаголовками ("к пролонгации" и т.д.)
C_WHITE  = "FFFFFF"

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(ws, row, col, value, bg=C_HEADER, bold=True, merge_end_col=None, merge_end_row=None):
    """Записывает стилизованную ячейку заголовка."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    cell.border    = thin_border()
    if merge_end_col or merge_end_row:
        er = merge_end_row or row
        ec = merge_end_col or col
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=er, end_column=ec
        )
    return cell

def dcell(ws, row, col, value):
    """Записывает ячейку с данными."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()
    return cell

def autowidth(ws, min_width=12):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None and not isinstance(c, MergedCell) else 0)
            for c in col
        )
        first = next(c for c in col if not isinstance(c, MergedCell))
        ws.column_dimensions[first.column_letter].width = max(max_len + 3, min_width)

# ─────────────────────────────────────────────────────────────────────────────
# Шаблонная двухстрочная шапка: Месяц/Менеджер | Пролонгации 1 (×3) | Пролонгации 2 (×3)
# ─────────────────────────────────────────────────────────────────────────────
def write_standard_header(ws, first_col_name):
    """
    Строки 1–2:
      A1:A2 = first_col_name (merged, rowspan 2)
      B1:D1 = "Пролонгации в первый месяц" (merged)
      E1:G1 = "Пролонгации через месяц"    (merged)
      B2 = "к пролонгации", C2 = "пролонгировано", D2 = "Коэффициент"
      E2 = "к пролонгации", F2 = "пролонгировано", G2 = "Коэффициент"
    """
    hcell(ws, 1, 1, first_col_name, bg=C_HEADER, merge_end_row=2)
    hcell(ws, 1, 2, "Пролонгации в первый месяц", bg=C_GROUP, merge_end_col=4)
    hcell(ws, 1, 5, "Пролонгации через месяц",    bg=C_GROUP, merge_end_col=7)
    for col, name in [(2, "к пролонгации"), (3, "пролонгировано"), (4, "Коэффициент"),
                      (5, "к пролонгации"), (6, "пролонгировано"), (7, "Коэффициент")]:
        hcell(ws, 2, col, name, bg=C_HEADER)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

def write_standard_data(ws, rows):
    """
    Записывает строки данных начиная с row 3.
    rows — список кортежей (label, к1, прод1, коэф1, к2, прод2, коэф2)
    """
    for i, (label, k1, p1, c1, k2, p2, c2) in enumerate(rows, start=3):
        dcell(ws, i, 1, label)
        dcell(ws, i, 2, fmt_money(k1))
        dcell(ws, i, 3, fmt_money(p1))
        dcell(ws, i, 4, fmt_pct(c1))
        dcell(ws, i, 5, fmt_money(k2))
        dcell(ws, i, 6, fmt_money(p2))
        dcell(ws, i, 7, fmt_pct(c2))

# ─────────────────────────────────────────────────────────────────────────────
# Создаём книгу с тремя листами
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)   # убираем дефолтный Sheet

# ── Лист 1: Весь отдел ───────────────────────────────────────────────────────
ws_dept = wb.create_sheet("Весь отдел")
write_standard_header(ws_dept, "Месяц")

dept_rows = [
    (row["Месяц"],
     row["к пролонгации 1"], row["пролонгировано 1"], row["Коэффициент 1"],
     row["к пролонгации 2"], row["пролонгировано 2"], row["Коэффициент 2"])
    for _, row in department.iterrows()
]
write_standard_data(ws_dept, dept_rows)
autowidth(ws_dept)

# ── Лист 2: Менеджеры за год ─────────────────────────────────────────────────
ws_year = wb.create_sheet("Менеджеры за год")
write_standard_header(ws_year, "Менеджер")

year_rows = [
    (row["Менеджер"],
     row["к пролонгации 1"], row["пролонгировано 1"], row["Коэффициент 1"],
     row["к пролонгации 2"], row["пролонгировано 2"], row["Коэффициент 2"])
    for _, row in year_manager.iterrows()
]
write_standard_data(ws_year, year_rows)
autowidth(ws_year)

# ── Лист 3: Менеджеры по месяцам ─────────────────────────────────────────────
# Две pivot-таблицы: Коэффициент 1 (сверху) и Коэффициент 2 (снизу)
ws_mon = wb.create_sheet("Менеджеры по месяцам")

month_cols_ordered = sorted(kpi_monthly["Месяц"].unique())   # '2023-01' .. '2023-12'
n_months = len(month_cols_ordered)
total_cols = 1 + n_months   # "Менеджер" + 12 месяцев

def write_pivot_block(ws, start_row, pivot_df, coef_col, block_title):
    """
    Записывает один блок pivot-таблицы.
    Строка start_row   : заголовок блока (merged по всей ширине)
    Строка start_row+1 : "Менеджер" | месяц1 | месяц2 | ...
    Строки далее       : данные
    """
    # Заголовок блока
    hcell(ws, start_row, 1, block_title, bg=C_GROUP,
          merge_end_col=total_cols)
    ws.row_dimensions[start_row].height = 22

    # Строка с подзаголовками
    hcell(ws, start_row + 1, 1, "Менеджер", bg=C_HEADER)
    for j, m in enumerate(month_cols_ordered, start=2):
        hcell(ws, start_row + 1, j, m, bg=C_HEADER)
    ws.row_dimensions[start_row + 1].height = 20

    # Данные
    managers = pivot_df["Менеджер"].tolist()
    for i, manager in enumerate(managers):
        row_n = start_row + 2 + i
        dcell(ws, row_n, 1, manager)
        for j, m in enumerate(month_cols_ordered, start=2):
            val = pivot_df.loc[pivot_df["Менеджер"] == manager, m].values
            pct = fmt_pct(val[0]) if len(val) > 0 and not pd.isna(val[0]) else "Нет данных"
            dcell(ws, row_n, j, pct)

    return start_row + 2 + len(managers)   # возвращаем следующую свободную строку

next_row = write_pivot_block(ws_mon, 1,  pivot_coef1, "Коэффициент 1", "Коэффициент 1")
next_row += 1   # пустая строка-разделитель
write_pivot_block(ws_mon, next_row, pivot_coef2, "Коэффициент 2", "Коэффициент 2")
autowidth(ws_mon, min_width=10)

# ── Лист 4: Итоги — сводка для руководителя ──────────────────────────────────
# Два блока на одном листе: отдел по месяцам + менеджеры за год.
# Место для графика оставляется ниже (пустые строки).
ws_sum = wb.create_sheet("Итоги", 0)   # первым листом

C_TITLE = "F4B942"   # оранжево-жёлтый для заголовков блоков

def write_summary_block(ws, start_row, title, first_col, rows_data):
    """
    Записывает один сводный блок:
      start_row   : строка с названием блока (merged на всю ширину = 7 столбцов)
      start_row+1 : двухстрочная шапка (такая же как на остальных листах)
      start_row+2+: данные
    Возвращает номер следующей свободной строки.
    """
    # Заголовок блока
    hcell(ws, start_row, 1, title, bg=C_TITLE, merge_end_col=7)
    ws.row_dimensions[start_row].height = 24

    # Стандартная двухстрочная шапка
    hcell(ws, start_row + 1, 1, first_col, bg=C_HEADER, merge_end_row=start_row + 2)
    hcell(ws, start_row + 1, 2, "Пролонгации в первый месяц", bg=C_GROUP, merge_end_col=4)
    hcell(ws, start_row + 1, 5, "Пролонгации через месяц",    bg=C_GROUP, merge_end_col=7)
    for col, name in [(2, "к пролонгации"), (3, "пролонгировано"), (4, "Коэффициент"),
                      (5, "к пролонгации"), (6, "пролонгировано"), (7, "Коэффициент")]:
        hcell(ws, start_row + 2, col, name, bg=C_HEADER)
    ws.row_dimensions[start_row + 1].height = 26
    ws.row_dimensions[start_row + 2].height = 26

    # Данные
    for i, (label, k1, p1, c1, k2, p2, c2) in enumerate(rows_data, start=start_row + 3):
        dcell(ws, i, 1, label)
        dcell(ws, i, 2, fmt_money(k1))
        dcell(ws, i, 3, fmt_money(p1))
        dcell(ws, i, 4, fmt_pct(c1))
        dcell(ws, i, 5, fmt_money(k2))
        dcell(ws, i, 6, fmt_money(p2))
        dcell(ws, i, 7, fmt_pct(c2))

    return start_row + 3 + len(rows_data)

# ── Блок 1: Весь отдел по месяцам + строка «Итого за 2023 год» ──────────────
next_r = write_summary_block(
    ws_sum, 1,
    "Весь отдел — по месяцам", "Месяц",
    [(r["Месяц"],
      r["к пролонгации 1"], r["пролонгировано 1"], r["Коэффициент 1"],
      r["к пролонгации 2"], r["пролонгировано 2"], r["Коэффициент 2"])
     for _, r in department.iterrows()]
)

# Строка «Итого за 2023 год» — жирная, тёмно-жёлтая
C_TOTAL = "FFD966"
dy = department_year
for col_idx, value in enumerate([
    "Итого за 2023 год",
    fmt_money(dy["к пролонгации 1"]),
    fmt_money(dy["пролонгировано 1"]),
    fmt_pct(dy["Коэффициент 1"]),
    fmt_money(dy["к пролонгации 2"]),
    fmt_money(dy["пролонгировано 2"]),
    fmt_pct(dy["Коэффициент 2"]),
], start=1):
    cell = ws_sum.cell(row=next_r, column=col_idx, value=value)
    cell.font      = Font(bold=True, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_TOTAL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()

# Разделитель между блоками
next_r += 2

# ── Блок 2: Менеджеры за год ──────────────────────────────────────────────────
next_r = write_summary_block(
    ws_sum, next_r,
    "Менеджеры — итог за 2023 год", "Менеджер",
    [(r["Менеджер"],
      r["к пролонгации 1"], r["пролонгировано 1"], r["Коэффициент 1"],
      r["к пролонгации 2"], r["пролонгировано 2"], r["Коэффициент 2"])
     for _, r in year_manager.iterrows()]
)

# Подсказка для вставки графика
next_r += 1
ws_sum.cell(row=next_r, column=1).value = "← Вставьте график здесь"
ws_sum.cell(row=next_r, column=1).font = Font(italic=True, color="999999", size=10)

autowidth(ws_sum)

# ─────────────────────────────────────────────────────────────────────────────
file_name = "prolongation_report.xlsx"
wb.save(file_name)
print(f"Отчёт сохранён: {file_name}")

# =============================================================================
# 8. КРАТКИЙ ВЫВОД В КОНСОЛЬ
# =============================================================================
print("\n── Весь отдел ──")
dept_print = department.copy()
dept_print["Коэффициент 1"] = dept_print["Коэффициент 1"].apply(fmt_pct)
dept_print["Коэффициент 2"] = dept_print["Коэффициент 2"].apply(fmt_pct)
print(dept_print[["Месяц","к пролонгации 1","пролонгировано 1","Коэффициент 1",
                   "к пролонгации 2","пролонгировано 2","Коэффициент 2"]].to_string(index=False))

print("\n── Менеджеры за год ──")
ym_print = year_manager.copy()
ym_print["Коэффициент 1"] = ym_print["Коэффициент 1"].apply(fmt_pct)
ym_print["Коэффициент 2"] = ym_print["Коэффициент 2"].apply(fmt_pct)
print(ym_print[["Менеджер","Коэффициент 1","Коэффициент 2"]].to_string(index=False))
